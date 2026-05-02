#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_yolo_dataset.py — SINGLE entry point.
   Download semua source data makanan Indonesia → 1 dataset siap pakai untuk
   YOLO/AI/ML training + nutrition output.

Cara pakai:
  python build_yolo_dataset.py

Output: dataset_makanan_indonesia.csv (~3000-3500 entri unik)

Kolom output:
  id              - Primary key 1..N
  food_name       - Nama makanan/bahan (label class untuk YOLO)
  category        - 1 dari 10 kategori
  image_url       - URL Wikipedia Commons (untuk download training image)
  calories_kcal   - Kalori (kkal)
  protein_g       - Protein (gram)
  fat_g           - Lemak (gram)
  carbohydrates_g - Karbohidrat (gram)
  sugar_g         - Gula (gram)
  sodium_mg       - Natrium (mg)
  fiber_g         - Serat (gram)
  source          - Provenance (mana data berasal — penting buat data scientist)

═══════════════════════════════════════════════════════════════════════════
SUMBER DATA YANG DI-MERGE (semua dari Indonesia)
═══════════════════════════════════════════════════════════════════════════

1. hand_curated (~2019 entri)
   Hand-curated manual: makanan jadi regional Indonesia dari 38 provinsi +
   bahan-bahan + bumbu + olahan (abon, tempe/tahu varian, sambal, kerupuk, dll).

2. TKPI_2020_Kemenkes (~1066 entri)
   Tabel Komposisi Pangan Indonesia 2020 RESMI Kementerian Kesehatan RI.
   Source: github.com/ancxlol/tkpi-2020-database
   Original PDF: Kemenkes RI publikasi resmi.

3. NutriSurvey_Indonesia (~960 entri)
   Database makanan Indonesia dari NutriSurvey (versi yang umum dipakai
   peneliti gizi di Indonesia, sebelum TKPI).
   Source: github.com/DiabetesLab-Project-Research/Indonesia.fta
   Original: nutrisurvey.de (program gizi standar Indonesia).
   Format: Excel (.xlsx) — butuh openpyxl. Kalau tidak ada, source ini di-skip.

═══════════════════════════════════════════════════════════════════════════
"""

import csv
import os
import sys
import urllib.request
import urllib.error

# ============================================================
# IMPORT BASE dataset dari generator hand-curated
# ============================================================
HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

try:
    from generate_indonesian_food_dataset import (
        BASE,
        find_image_pool,
        pick_image,
        IMAGE_POOL,
        wc,
    )
except ImportError as e:
    print(f"FATAL: tidak bisa import generate_indonesian_food_dataset.py")
    print(f"       Pastikan file ini ada di folder yang sama.")
    print(f"       Error: {e}")
    sys.exit(1)


# ============================================================
# SOURCE LIST — semua data makanan/gizi Indonesia
# ============================================================
TKPI_URL = "https://raw.githubusercontent.com/ancxlol/tkpi-2020-database/main/data/tkpi_2020_pages_15_83.csv"
TKPI_LOCAL = os.path.join(HERE, "_cache_tkpi_2020.csv")

NUTRISURVEY_URL = "https://github.com/DiabetesLab-Project-Research/Indonesia.fta/raw/master/Indonesia.fta.xlsx"
NUTRISURVEY_LOCAL = os.path.join(HERE, "_cache_nutrisurvey_indonesia.xlsx")


def download_file(url, local_path, label):
    """Download URL ke local_path. Return True kalau berhasil/sudah ada."""
    if os.path.exists(local_path) and os.path.getsize(local_path) > 1000:
        print(f"  ✓ {label} cache sudah ada ({os.path.getsize(local_path):,} bytes)")
        return True
    print(f"  Download {label}...")
    print(f"    URL: {url}")
    try:
        req = urllib.request.Request(
            url,
            headers={"User-Agent": "Mozilla/5.0 (yolo-dataset-builder/1.0)"},
        )
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = resp.read()
        with open(local_path, "wb") as f:
            f.write(data)
        print(f"    ✓ Saved {len(data):,} bytes → {local_path}")
        return True
    except (urllib.error.URLError, urllib.error.HTTPError, Exception) as e:
        print(f"    ✗ Gagal: {e}")
        return False


# ============================================================
# CATEGORY MAPPING (TKPI/NutriSurvey code → kategori dataset)
# ============================================================
TKPI_CAT_MAP = {
    "A": "umbi",         # serealia (beras, jagung, gandum)
    "B": "umbi",         # umbi (singkong, ubi, kentang)
    "C": "kacang",       # kacang & olahan
    "D": "sayur",        # sayuran
    "E": "buah",         # buah-buahan
    "F": "daging",       # daging
    "G": "ikan",         # ikan/seafood (split via name keyword)
    "H": "ayam",         # telur
    "I": "kacang",       # susu (kelompok kacang)
    "J": "bumbu",        # lemak/minyak
    "K": "bumbu",        # gula, sirup
    "L": "bumbu",        # bumbu/rempah
    "M": "makanan-jadi", # minuman
    "N": "makanan-jadi", # olahan komposit
}


def categorize(name, category_code):
    """Convert category code → kategori dataset."""
    cat = TKPI_CAT_MAP.get(category_code, "makanan-jadi")
    name_lower = name.lower()
    if cat == "ikan":
        seafood_kw = ["udang", "cumi", "sotong", "gurita", "kepiting", "rajungan",
                      "lobster", "kerang", "tiram", "remis", "rumput laut",
                      "teripang", "ebi", "simping"]
        if any(w in name_lower for w in seafood_kw):
            return "seafood"
    return cat


def categorize_by_name(name):
    """Heuristic categorization based on name keywords (fallback untuk
    source tanpa category code)."""
    n = name.lower()
    # Order matters - lebih spesifik dulu
    if any(w in n for w in ["nasi ", "soto", "sate", "mie ", "bakmi", "bakso",
                              "rendang", "dendeng", "gulai", "kari ", "rawon",
                              "semur", "opor", "pempek", "tekwan", "siomay",
                              "batagor", "martabak", "klepon", "lemper",
                              "lumpia", "kue ", "bakwan", "tahu isi", "cireng",
                              "cilok", "es ", "wedang", "jamu", "pecel",
                              "gado", "rujak", "lotek", "ketoprak", "asinan",
                              "dadar gulung", "onde-onde", "putu", "sambal "]):
        return "makanan-jadi"
    if any(w in n for w in ["ikan ", "teri", "bandeng", "tuna", "tongkol",
                              "lele", "nila", "patin", "gurame", "kakap",
                              "kerapu", "mujair", "bilis", "sardin"]):
        return "ikan"
    if any(w in n for w in ["udang", "cumi", "sotong", "kepiting", "rajungan",
                              "kerang", "tiram", "lobster"]):
        return "seafood"
    if any(w in n for w in ["daging sapi", "daging kambing", "daging domba",
                              "daging babi", "hati ", "iga ", "buntut", "lidah",
                              "otak ", "babat"]):
        return "daging"
    if any(w in n for w in ["ayam", "bebek", "itik", "puyuh", "telur", "kalkun"]):
        return "ayam"
    if any(w in n for w in ["mangga", "pisang", "pepaya", "nanas", "jambu",
                              "rambutan", "durian", "salak", "manggis",
                              "jeruk", "anggur", "apel", "alpukat", "kelengkeng",
                              "leci", "duku", "langsat", "sirsak", "nangka"]):
        return "buah"
    if any(w in n for w in ["bayam", "kangkung", "sawi", "kol ", "wortel",
                              "kentang", "tomat", "terong", "timun", "buncis",
                              "tauge", "jagung", "daun ", "pakis", "rebung",
                              "bunga ", "kembang ", "selada", "pokcoy"]):
        return "sayur"
    if any(w in n for w in ["kacang", "kedelai", "tempe", "tahu", "oncom",
                              "kemiri", "kenari", "mete", "wijen"]):
        return "kacang"
    if any(w in n for w in ["jahe", "kunyit", "kencur", "lengkuas", "serai",
                              "ketumbar", "merica", "lada", "asam", "cengkeh",
                              "pala", "kayu manis", "gula", "garam", "minyak",
                              "santan", "kecap", "saus", "petis", "terasi",
                              "ebi", "bawang"]):
        return "bumbu"
    if any(w in n for w in ["singkong", "ubi", "talas", "tepung", "sagu",
                              "tapioka", "bengkuang", "lobak", "porang"]):
        return "umbi"
    return "makanan-jadi"


def safe_float(s):
    """Parse string ke float, handle '-', 'tr', empty, koma decimal."""
    if s is None:
        return 0.0
    s = str(s).strip()
    if not s or s in ("-", "tr", "trace", "nan", "NaN", "null", "None"):
        return 0.0
    try:
        return float(s.replace(",", "."))
    except (ValueError, TypeError):
        return 0.0


# ============================================================
# PARSER 1: TKPI 2020 (CSV)
# ============================================================
def parse_tkpi():
    """Parse TKPI 2020 CSV → list of unified dicts."""
    if not os.path.exists(TKPI_LOCAL):
        return []
    out = []
    try:
        with open(TKPI_LOCAL, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = " ".join(((row.get("name") or "").strip()).lower().split())
                if not name or len(name) < 2:
                    continue
                tkpi_cat = (row.get("category") or "").strip().upper()[:1]
                category = categorize(name, tkpi_cat)
                cal  = safe_float(row.get("energy_kcal"))
                pro  = safe_float(row.get("protein_g"))
                fat  = safe_float(row.get("fat_g"))
                carb = safe_float(row.get("carb_g"))
                fib  = safe_float(row.get("fiber_g"))
                sod  = safe_float(row.get("sodium_mg"))
                sug  = round(carb * 0.7, 1) if category == "buah" else 0.0
                out.append({
                    "name": name, "category": category,
                    "cal": cal, "pro": pro, "fat": fat, "carb": carb,
                    "sug": sug, "sod": sod, "fib": fib,
                })
    except Exception as e:
        print(f"    ! Parse TKPI gagal: {e}")
        return []
    return out


# ============================================================
# PARSER 2: NutriSurvey Indonesia.fta (Excel)
# ============================================================
def parse_nutrisurvey():
    """Parse Indonesia.fta.xlsx → list of unified dicts.

    Butuh openpyxl. Kalau tidak ada, return [].
    """
    if not os.path.exists(NUTRISURVEY_LOCAL):
        return []
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(f"    ! openpyxl tidak terinstall.")
        print(f"      Install: pip install openpyxl")
        print(f"      → NutriSurvey Indonesia di-skip")
        return []

    out = []
    try:
        wb = load_workbook(NUTRISURVEY_LOCAL, read_only=True, data_only=True)
        # Cari sheet pertama atau yang punya data
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Cari header row
            header_row = None
            header_cells = None
            for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
                if not row:
                    continue
                row_lower = [str(c or "").lower() for c in row]
                if any("name" in c or "nama" in c for c in row_lower):
                    if any("energy" in c or "kalori" in c or "energi" in c for c in row_lower):
                        header_row = row_idx
                        header_cells = row_lower
                        break
            if header_row is None:
                continue

            # Map column index
            def find_col(*keywords):
                for i, c in enumerate(header_cells):
                    for kw in keywords:
                        if kw in c:
                            return i
                return None

            col_name = find_col("name", "nama")
            col_cal  = find_col("energy", "kalori", "energi", "kcal")
            col_pro  = find_col("protein")
            col_fat  = find_col("fat", "lemak")
            col_carb = find_col("carbo", "karbohidrat", "karb")
            col_fib  = find_col("fiber", "serat")
            col_sod  = find_col("sodium", "natrium", "na ")

            if col_name is None or col_cal is None:
                continue

            for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
                if not row or col_name >= len(row):
                    continue
                name_raw = row[col_name]
                if not name_raw:
                    continue
                name = " ".join(str(name_raw).strip().lower().split())
                if len(name) < 2:
                    continue
                cal  = safe_float(row[col_cal])  if col_cal  is not None and col_cal  < len(row) else 0.0
                pro  = safe_float(row[col_pro])  if col_pro  is not None and col_pro  < len(row) else 0.0
                fat  = safe_float(row[col_fat])  if col_fat  is not None and col_fat  < len(row) else 0.0
                carb = safe_float(row[col_carb]) if col_carb is not None and col_carb < len(row) else 0.0
                fib  = safe_float(row[col_fib])  if col_fib  is not None and col_fib  < len(row) else 0.0
                sod  = safe_float(row[col_sod])  if col_sod  is not None and col_sod  < len(row) else 0.0
                category = categorize_by_name(name)
                sug = round(carb * 0.7, 1) if category == "buah" else 0.0
                out.append({
                    "name": name, "category": category,
                    "cal": cal, "pro": pro, "fat": fat, "carb": carb,
                    "sug": sug, "sod": sod, "fib": fib,
                })
            break  # cukup parse sheet pertama yang valid
    except Exception as e:
        print(f"    ! Parse NutriSurvey gagal: {e}")
        return []
    return out


# ============================================================
# Image router (re-use dari generator)
# ============================================================
def get_image_url(food_name, category):
    pool = find_image_pool(food_name, category)
    return pick_image(pool, 0)


# ============================================================
# Helpers
# ============================================================
def normalize_name(s):
    return " ".join(s.lower().strip().split())


# ============================================================
# MAIN
# ============================================================
def main():
    print("=" * 76)
    print("BUILD DATASET MAKANAN INDONESIA — multi-source merge untuk YOLO + nutrition")
    print("=" * 76)
    print()

    # --- Step 1: Hand-curated dataset (built-in) ---
    print(f"[1/4] Hand-curated dataset (built-in)...")
    print(f"      {len(BASE)} entri (regional foods + bahan + olahan)")

    # --- Step 2: Download TKPI 2020 (Kemenkes RI) ---
    print(f"\n[2/4] TKPI 2020 (Kemenkes RI)...")
    tkpi_ok = download_file(TKPI_URL, TKPI_LOCAL, "TKPI 2020")
    tkpi_rows = parse_tkpi() if tkpi_ok else []
    print(f"      {len(tkpi_rows)} entri parsed")

    # --- Step 3: Download NutriSurvey Indonesia ---
    print(f"\n[3/4] NutriSurvey Indonesia (Indonesia.fta)...")
    ns_ok = download_file(NUTRISURVEY_URL, NUTRISURVEY_LOCAL, "NutriSurvey Indonesia")
    ns_rows = parse_nutrisurvey() if ns_ok else []
    print(f"      {len(ns_rows)} entri parsed")

    # --- Step 4: Merge & dedupe ---
    print(f"\n[4/4] Merge & dedupe by food_name...")
    seen_names = set()
    rows_out = []
    rid = 1
    skipped_no_nutrition = 0

    # Hand-curated dulu (priority — ada image specific)
    for item in BASE:
        food_name, cat, cal, pro, fat, carb, sug, sod, fib, img_file = item
        norm = normalize_name(food_name)
        if norm in seen_names:
            continue
        if cal == 0 and pro == 0 and fat == 0 and carb == 0:
            skipped_no_nutrition += 1
            continue
        seen_names.add(norm)
        if img_file:
            img_url = wc(img_file)
        else:
            img_url = get_image_url(food_name, cat)
        rows_out.append({
            "id":               rid,
            "food_name":        food_name,
            "category":         cat,
            "image_url":        img_url,
            "calories_kcal":    cal,
            "protein_g":        pro,
            "fat_g":            fat,
            "carbohydrates_g":  carb,
            "sugar_g":          sug,
            "sodium_mg":        sod,
            "fiber_g":          fib,
            "source":           "hand_curated",
        })
        rid += 1

    # TKPI 2020
    tkpi_added = 0
    for t in tkpi_rows:
        norm = normalize_name(t["name"])
        if norm in seen_names:
            continue
        if t["cal"] == 0 and t["pro"] == 0 and t["fat"] == 0 and t["carb"] == 0:
            skipped_no_nutrition += 1
            continue
        seen_names.add(norm)
        rows_out.append({
            "id":               rid,
            "food_name":        t["name"],
            "category":         t["category"],
            "image_url":        get_image_url(t["name"], t["category"]),
            "calories_kcal":    t["cal"],
            "protein_g":        t["pro"],
            "fat_g":            t["fat"],
            "carbohydrates_g":  t["carb"],
            "sugar_g":          t["sug"],
            "sodium_mg":        t["sod"],
            "fiber_g":          t["fib"],
            "source":           "TKPI_2020_Kemenkes",
        })
        rid += 1
        tkpi_added += 1

    # NutriSurvey
    ns_added = 0
    for n in ns_rows:
        norm = normalize_name(n["name"])
        if norm in seen_names:
            continue
        if n["cal"] == 0 and n["pro"] == 0 and n["fat"] == 0 and n["carb"] == 0:
            skipped_no_nutrition += 1
            continue
        seen_names.add(norm)
        rows_out.append({
            "id":               rid,
            "food_name":        n["name"],
            "category":         n["category"],
            "image_url":        get_image_url(n["name"], n["category"]),
            "calories_kcal":    n["cal"],
            "protein_g":        n["pro"],
            "fat_g":            n["fat"],
            "carbohydrates_g":  n["carb"],
            "sugar_g":          n["sug"],
            "sodium_mg":        n["sod"],
            "fiber_g":          n["fib"],
            "source":           "NutriSurvey_Indonesia",
        })
        rid += 1
        ns_added += 1

    print(f"      {len(BASE)} hand_curated + {tkpi_added} TKPI + {ns_added} NutriSurvey = {len(rows_out)} unique")

    # --- Write final CSV ---
    out_path = os.path.join(HERE, "dataset_makanan_indonesia.csv")
    fieldnames = [
        "id", "food_name", "category", "image_url",
        "calories_kcal", "protein_g", "fat_g", "carbohydrates_g",
        "sugar_g", "sodium_mg", "fiber_g",
        "source",
    ]
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows_out)

    # --- Summary ---
    cat_counts = {}
    src_counts = {}
    for r in rows_out:
        cat_counts[r["category"]] = cat_counts.get(r["category"], 0) + 1
        src_counts[r["source"]] = src_counts.get(r["source"], 0) + 1

    print(f"\n" + "=" * 76)
    print(f"DONE")
    print(f"=" * 76)
    print(f"\nOutput: {out_path}")
    print(f"Total: {len(rows_out)} entri unik (deduped by food_name)")
    if skipped_no_nutrition:
        print(f"Skipped: {skipped_no_nutrition} entri (nutrisi 0 semua)")
    print()
    print(f"Per kategori:")
    for cat, n in sorted(cat_counts.items()):
        print(f"  {cat:14s}  {n:5d}")
    print()
    print(f"Per source (provenance):")
    for src, n in sorted(src_counts.items(), key=lambda x: -x[1]):
        print(f"  {src:25s}  {n:5d}")
    print()
    print(f"Schema (siap untuk YOLO + nutrition output):")
    print(f"  food_name        → label class untuk YOLO")
    print(f"  image_url        → download untuk training data")
    print(f"  calories_kcal..  → 7 kolom nutrisi untuk output ke user saat detect")
    print(f"  source           → provenance (penting untuk data scientist cleaning)")
    print()
    print(f"Next steps untuk data scientist:")
    print(f"  1. Validasi: cek nutrisi outliers, drop entri yang aneh")
    print(f"  2. Image cleanup: download image_url, drop yang 404, anotasi bbox")
    print(f"  3. Train YOLO classifier (food_name = label class)")
    print(f"  4. Inference: detect food → lookup CSV by food_name → output 7 nutrisi")


if __name__ == "__main__":
    main()
